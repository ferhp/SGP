# Generación de exportaciones (PDF y XLSX) para las tablas de cada vista.
#
# Las filas llegan ya formateadas como texto (fechas, montos, porcentajes)
# para que ambos formatos muestren exactamente lo mismo que la tabla en
# pantalla.
from datetime import datetime
from io import BytesIO

from flask import abort, send_file
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

AZUL = "1A73E8"


def moneda(valor):
    try:
        return f"${valor:,.2f}"
    except (TypeError, ValueError):
        return "$0.00"


def respuesta_exportacion(formato, nombre, titulo, columnas, filas):
    """Regresa la respuesta Flask con el archivo pedido (xlsx o pdf)."""
    fecha = datetime.now().strftime("%Y-%m-%d")
    if formato == "xlsx":
        return _xlsx(f"{nombre}_{fecha}.xlsx", titulo, columnas, filas)
    if formato == "pdf":
        return _pdf(f"{nombre}_{fecha}.pdf", titulo, columnas, filas)
    abort(404)


def _xlsx(nombre_archivo, titulo, columnas, filas):
    libro = Workbook()
    hoja = libro.active
    hoja.title = titulo[:31] or "Datos"

    hoja.append([titulo])
    hoja["A1"].font = Font(bold=True, size=14, color=AZUL)
    hoja.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columnas))

    hoja.append(list(columnas))
    for celda in hoja[2]:
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = PatternFill("solid", fgColor=AZUL)
        celda.alignment = Alignment(vertical="center")

    for fila in filas:
        hoja.append(["" if v is None else v for v in fila])

    for i, columna in enumerate(columnas, start=1):
        ancho = len(str(columna))
        for fila in filas:
            valor = fila[i - 1]
            if valor is not None:
                ancho = max(ancho, len(str(valor)))
        hoja.column_dimensions[get_column_letter(i)].width = min(45, ancho + 3)
    hoja.freeze_panes = "A3"

    contenido = BytesIO()
    libro.save(contenido)
    contenido.seek(0)
    return send_file(
        contenido, as_attachment=True, download_name=nombre_archivo,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def _pdf(nombre_archivo, titulo, columnas, filas):
    contenido = BytesIO()
    documento = SimpleDocTemplate(
        contenido, pagesize=landscape(letter), title=titulo,
        leftMargin=12 * mm, rightMargin=12 * mm,
        topMargin=12 * mm, bottomMargin=12 * mm,
    )
    estilos = getSampleStyleSheet()
    estilo_titulo = ParagraphStyle("titulo", parent=estilos["Title"],
                                   textColor=colors.HexColor("#1A73E8"),
                                   fontSize=15, spaceAfter=4)
    estilo_celda = ParagraphStyle("celda", parent=estilos["Normal"],
                                  fontSize=8, leading=10)
    estilo_encabezado = ParagraphStyle("encabezado", parent=estilo_celda,
                                       textColor=colors.white, fontName="Helvetica-Bold")

    datos = [[Paragraph(str(c), estilo_encabezado) for c in columnas]]
    for fila in filas:
        datos.append([Paragraph("" if v is None else str(v), estilo_celda)
                      for v in fila])

    tabla = Table(datos, repeatRows=1)
    tabla.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1A73E8")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1),
         [colors.white, colors.HexColor("#F1F3F4")]),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#DADCE0")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]))

    generado = datetime.now().strftime("%d/%m/%Y %H:%M")
    documento.build([
        Paragraph(titulo, estilo_titulo),
        Paragraph(f"SIGEP · Generado el {generado}", estilos["Normal"]),
        Spacer(1, 6 * mm),
        tabla,
    ])
    contenido.seek(0)
    return send_file(contenido, as_attachment=True,
                     download_name=nombre_archivo, mimetype="application/pdf")
